from __future__ import annotations

import os
from typing import Any, Dict, List, Optional
import csv
from datetime import datetime, timezone

# Try to load a simple .env file from the repository root if present and
# environment variables required for MSSQL are not set. This helps when the
# developer added vars to ~/.zshrc but the server process wasn't restarted.
def _load_dotenv_if_missing():
    needed = ["MSSQL_HOST", "MSSQL_DB", "MSSQL_USER", "MSSQL_PASSWORD"]
    missing = [k for k in needed if not os.environ.get(k)]
    if not missing:
        return
    # Look for .env at repo root or .env in cwd
    candidates = [".env", ".env.local", ".env.development"]
    root = Path(__file__).resolve().parents[2]
    for name in candidates:
        p = root / name
        if p.exists():
            try:
                for ln in p.read_text(encoding='utf-8').splitlines():
                    ln = ln.strip()
                    if not ln or ln.startswith('#') or '=' not in ln:
                        continue
                    k, v = ln.split('=', 1)
                    k = k.strip()
                    v = v.strip().strip('"').strip("'")
                    # do not overwrite existing env
                    if not os.environ.get(k):
                        os.environ[k] = v
                break
            except Exception:
                continue

    # As a last resort, try parsing common shell rc file (~/.zshrc) for exported MSSQL_* vars
    if any(k not in os.environ for k in needed):
        try:
            home_rc = Path.home() / '.zshrc'
            if home_rc.exists():
                for ln in home_rc.read_text(encoding='utf-8').splitlines():
                    ln = ln.strip()
                    if not ln or not ln.startswith('export '):
                        continue
                    try:
                        rest = ln[len('export '):]
                        if '=' not in rest:
                            continue
                        k, v = rest.split('=', 1)
                        k = k.strip()
                        v = v.strip().strip('"').strip("'")
                        if k in needed and not os.environ.get(k):
                            os.environ[k] = v
                    except Exception:
                        continue
        except Exception:
            pass

# Run once at import time to populate environment if possible
from pathlib import Path
_load_dotenv_if_missing()


def _yn(val: Optional[str], default_yes: bool = True) -> str:
    v = (val or ("yes" if default_yes else "no")).strip().lower()
    if v in {"1", "true", "t", "y", "yes"}:
        return "yes"
    if v in {"0", "false", "f", "n", "no"}:
        return "no"
    # pass-through for values like 'strict'
    return v


def sqlalchemy_url_from_env() -> Optional[str]:
    host = os.environ.get("MSSQL_HOST")
    db = os.environ.get("MSSQL_DB")
    user = os.environ.get("MSSQL_USER")
    pwd = os.environ.get("MSSQL_PASSWORD")
    port = os.environ.get("MSSQL_PORT", "1433")
    encrypt = _yn(os.environ.get("MSSQL_ENCRYPT", "yes"), default_yes=True)
    trust = _yn(os.environ.get("MSSQL_TRUST_CERT", "yes" if encrypt == "yes" else "no"))
    if not (host and db and user and pwd):
        return None
    driver = os.environ.get("ODBC_DRIVER", "ODBC Driver 18 for SQL Server")
    return (
        f"mssql+pyodbc://{user}:{pwd}@{host},{port}/{db}?driver={driver.replace(' ', '+')}"
        f"&Encrypt={encrypt}&TrustServerCertificate={trust}"
    )


def create_engine_lazy(url: str):
    try:
        import sqlalchemy as sa  # type: ignore
    except Exception as e:
        raise ImportError(
            "SQLAlchemy is required for MSSQL operations. Install `sqlalchemy` and `pyodbc`"
        ) from e
    return sa.create_engine(url, pool_pre_ping=True)


def run_query_to_dicts(sql: str, max_rows: Optional[int] = None) -> List[Dict[str, Any]]:
    url = sqlalchemy_url_from_env()
    if not url:
        raise RuntimeError(
            "MSSQL environment variables not set. Set MSSQL_HOST, MSSQL_DB, MSSQL_USER, MSSQL_PASSWORD."
        )
    engine = create_engine_lazy(url)
    rows: List[Dict[str, Any]] = []
    with engine.connect() as conn:
        result = conn.execute(sql)  # type: ignore[arg-type]
        cols = result.keys()
        for i, r in enumerate(result):
            rows.append({k: r[idx] for idx, k in enumerate(cols)})
            if max_rows is not None and i + 1 >= max_rows:
                break
    return rows


def ensure_hackathon_schema() -> Dict[str, Any]:
    url = sqlalchemy_url_from_env()
    if not url:
        raise RuntimeError(
            "MSSQL environment variables not set. Set MSSQL_HOST, MSSQL_DB, MSSQL_USER, MSSQL_PASSWORD."
        )
    engine = create_engine_lazy(url)
    with engine.begin() as conn:
        # Employees
        conn.exec_driver_sql(
            """
            IF OBJECT_ID('dbo.ht_Employees', 'U') IS NULL
            CREATE TABLE dbo.ht_Employees (
                employee_id NVARCHAR(50) NOT NULL PRIMARY KEY,
                name NVARCHAR(200) NULL,
                department NVARCHAR(200) NULL,
                city NVARCHAR(100) NULL
            )
            """
        )
        # Transactions
        conn.exec_driver_sql(
            """
            IF OBJECT_ID('dbo.ht_Transactions', 'U') IS NULL
            CREATE TABLE dbo.ht_Transactions (
                txn_id NVARCHAR(50) NOT NULL PRIMARY KEY,
                employee_id NVARCHAR(50) NULL,
                merchant NVARCHAR(200) NULL,
                city NVARCHAR(100) NULL,
                category NVARCHAR(100) NULL,
                amount DECIMAL(12,2) NULL,
                [timestamp] DATETIME2 NULL,
                channel NVARCHAR(50) NULL,
                card_id NVARCHAR(50) NULL
            )
            """
        )
        # Models
        conn.exec_driver_sql(
            """
            IF OBJECT_ID('dbo.ht_Models', 'U') IS NULL
            CREATE TABLE dbo.ht_Models (
                model_id INT IDENTITY(1,1) NOT NULL PRIMARY KEY,
                algo NVARCHAR(50) NOT NULL,
                created_at DATETIME2 NOT NULL DEFAULT SYSUTCDATETIME(),
                metrics_json NVARCHAR(MAX) NULL
            )
            """
        )
        # Scores
        conn.exec_driver_sql(
            """
            IF OBJECT_ID('dbo.ht_Scores', 'U') IS NULL
            CREATE TABLE dbo.ht_Scores (
                score_id INT IDENTITY(1,1) NOT NULL PRIMARY KEY,
                txn_id NVARCHAR(50) NULL,
                model_id INT NULL,
                fraud_score FLOAT NULL,
                compliant BIT NULL,
                reason NVARCHAR(4000) NULL
            )
            """
        )
        # Logs (preferred table)
        conn.exec_driver_sql(
            """
            IF OBJECT_ID('dbo.ht_AppsLogs', 'U') IS NULL
            CREATE TABLE dbo.ht_AppsLogs (
                id INT IDENTITY(1,1) PRIMARY KEY,
                ts DATETIME2 NOT NULL,
                event_type NVARCHAR(100) NOT NULL,
                payload NVARCHAR(MAX) NOT NULL
            )
            """
        )
        # Optional migration: copy from legacy ht_AppLogs if exists and destination empty
        conn.exec_driver_sql(
            """
            IF OBJECT_ID('dbo.ht_AppLogs', 'U') IS NOT NULL AND NOT EXISTS (SELECT 1 FROM dbo.ht_AppsLogs)
            INSERT INTO dbo.ht_AppsLogs(ts, event_type, payload)
            SELECT ts, event_type, payload FROM dbo.ht_AppLogs
            """
        )
        # Indexes
        conn.exec_driver_sql(
            """
            IF NOT EXISTS (SELECT name FROM sys.indexes WHERE name = 'IX_ht_Transactions_Employee_Timestamp')
            CREATE INDEX IX_ht_Transactions_Employee_Timestamp ON dbo.ht_Transactions(employee_id, [timestamp]);
            IF NOT EXISTS (SELECT name FROM sys.indexes WHERE name = 'IX_ht_Transactions_Merchant_Timestamp')
            CREATE INDEX IX_ht_Transactions_Merchant_Timestamp ON dbo.ht_Transactions(merchant, [timestamp]);
            """
        )
    return {"status": "ok", "message": "Hackathon schema ensured (ht_*)"}


def _parse_ts(value: str):
    if not value:
        return None
    try:
        # Accept ISO-8601 with timezone Z or offset
        v = value.replace("Z", "+00:00")
        dt = datetime.fromisoformat(v)
        # Convert to naive UTC for DATETIME2 compatibility
        if dt.tzinfo is not None:
            dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
        return dt
    except Exception:
        try:
            # Fallback common format
            return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
        except Exception:
            return None


def load_transactions_csv(path: str, truncate: bool = False, limit: Optional[int] = None, batch_size: int = 100) -> Dict[str, Any]:
    url = sqlalchemy_url_from_env()
    if not url:
        raise RuntimeError(
            "MSSQL environment variables not set. Set MSSQL_HOST, MSSQL_DB, MSSQL_USER, MSSQL_PASSWORD."
        )
    engine = create_engine_lazy(url)
    inserted = 0
    insert_sql = """
                        INSERT INTO dbo.ht_Transactions (
                            txn_id, employee_id, merchant, city, category, amount, [timestamp], channel, card_id
                        ) VALUES (?,?,?,?,?,?,?,?,?)
                        """
    with engine.begin() as conn:
        if truncate:
            conn.exec_driver_sql("TRUNCATE TABLE dbo.ht_Transactions")
        with open(path, newline="") as f:
            reader = csv.DictReader(f)
            batch: List[tuple] = []
            for i, row in enumerate(reader):
                if limit is not None and i >= limit:
                    break
                try:
                    params = (
                        row.get("txn_id"),
                        row.get("employee_id"),
                        row.get("merchant"),
                        row.get("city"),
                        row.get("category"),
                        float(row.get("amount") or 0.0),
                        _parse_ts(row.get("timestamp") or ""),
                        row.get("channel"),
                        row.get("card_id"),
                    )
                    batch.append(params)
                    inserted += 1
                    if batch_size and len(batch) >= int(batch_size):
                        conn.exec_driver_sql(insert_sql, batch)
                        batch = []
                except Exception:
                    # Ignore duplicates or bad rows for hackathon speed
                    continue
            if batch:
                conn.exec_driver_sql(insert_sql, batch)
    return {"status": "ok", "inserted": inserted}


def _normalize_header(name: Optional[str]) -> str:
    if not name:
        return ""
    return (
        str(name)
        .strip()
        .lower()
        .replace(" ", "_")
        .replace("-", "_")
    )


def load_transactions_excel(
    path: str, sheet: Optional[str] = None, truncate: bool = False, limit: Optional[int] = None, batch_size: int = 100
) -> Dict[str, Any]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise ImportError(
            "openpyxl is required for Excel ingestion. Install with: pip install openpyxl"
        ) from e

    url = sqlalchemy_url_from_env()
    if not url:
        raise RuntimeError(
            "MSSQL environment variables not set. Set MSSQL_HOST, MSSQL_DB, MSSQL_USER, MSSQL_PASSWORD."
        )
    engine = create_engine_lazy(url)

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active

    rows_iter = ws.iter_rows(min_row=1, values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        return {"status": "ok", "inserted": 0}
    norm_headers = [_normalize_header(h) for h in headers]
    # Map normalized headers to canonical columns
    field_map: Dict[str, Optional[int]] = {k: None for k in [
        "txn_id", "employee_id", "merchant", "city", "category", "amount", "timestamp", "channel", "card_id"
    ]}
    aliases = {
        "txn_id": {"txn_id", "transaction_id", "id"},
        "employee_id": {"employee_id", "emp_id", "employee"},
        "merchant": {"merchant", "vendor"},
        "city": {"city", "location_city"},
        "category": {"category", "cat"},
        "amount": {"amount", "amt", "total", "value"},
        "timestamp": {"timestamp", "time", "date", "datetime"},
        "channel": {"channel", "payment_channel", "method"},
        "card_id": {"card_id", "card", "card_number"},
    }
    for idx, nh in enumerate(norm_headers):
        for canon, names in aliases.items():
            if nh in names and field_map[canon] is None:
                field_map[canon] = idx
                break

    inserted = 0
    insert_sql = """
                    INSERT INTO dbo.ht_Transactions (
                        txn_id, employee_id, merchant, city, category, amount, [timestamp], channel, card_id
                    ) VALUES (?,?,?,?,?,?,?,?,?)
                    """
    with engine.begin() as conn:
        if truncate:
            conn.exec_driver_sql("TRUNCATE TABLE dbo.ht_Transactions")
        batch: List[tuple] = []
        for i, row in enumerate(rows_iter):
            if limit is not None and i >= limit:
                break
            def get(col: str):
                j = field_map.get(col)
                if j is None:
                    return None
                return row[j]

            ts_val = get("timestamp")
            ts_dt = None
            if isinstance(ts_val, datetime):
                ts_dt = ts_val
            elif isinstance(ts_val, str):
                ts_dt = _parse_ts(ts_val)

            amt_val = get("amount")
            try:
                amount_f = float(amt_val) if amt_val is not None else 0.0
            except Exception:
                amount_f = 0.0

            try:
                params = (
                    get("txn_id"),
                    get("employee_id"),
                    get("merchant"),
                    get("city"),
                    get("category"),
                    amount_f,
                    ts_dt,
                    get("channel"),
                    get("card_id"),
                )
                batch.append(params)
                inserted += 1
                if batch_size and len(batch) >= int(batch_size):
                    conn.exec_driver_sql(insert_sql, batch)
                    batch = []
            except Exception:
                # Skip duplicates/bad rows
                continue
        if batch:
            conn.exec_driver_sql(insert_sql, batch)
    return {"status": "ok", "inserted": inserted}


def truncate_transactions() -> Dict[str, Any]:
    url = sqlalchemy_url_from_env()
    if not url:
        raise RuntimeError(
            "MSSQL environment variables not set. Set MSSQL_HOST, MSSQL_DB, MSSQL_USER, MSSQL_PASSWORD."
        )
    engine = create_engine_lazy(url)
    with engine.begin() as conn:
        conn.exec_driver_sql("TRUNCATE TABLE dbo.ht_Transactions")
    return {"status": "ok", "message": "dbo.ht_Transactions truncated"}


def query_transactions(
    employee_id: Optional[list[str]] | None = None,
    merchant: Optional[list[str]] | None = None,
    city: Optional[list[str]] | None = None,
    category: Optional[list[str]] | None = None,
    channel: Optional[list[str]] | None = None,
    card_id: Optional[list[str]] | None = None,
    min_amount: Optional[float] = None,
    max_amount: Optional[float] = None,
    start_ts: Optional[str] = None,
    end_ts: Optional[str] = None,
    page: int = 0,
    page_size: int = 50,
    sort_by: Optional[str] = None,
    sort_dir: str = "desc",
) -> Dict[str, Any]:
    url = sqlalchemy_url_from_env()
    if not url:
        raise RuntimeError(
            "MSSQL environment variables not set. Set MSSQL_HOST, MSSQL_DB, MSSQL_USER, MSSQL_PASSWORD."
        )
    engine = create_engine_lazy(url)

    where = []
    params: List[Any] = []
    like = lambda v: f"%{v}%"
    def _multi_field(name: str, values: Optional[list[str]]):
        if not values:
            return
        # build OR clauses for LIKE matching each value
        clauses = []
        for _ in values:
            clauses.append(f"{name} LIKE ?")
        where.append("(" + " OR ".join(clauses) + ")")
        for v in values:
            params.append(like(v))

    _multi_field('employee_id', employee_id)
    _multi_field('merchant', merchant)
    _multi_field('city', city)
    _multi_field('category', category)
    _multi_field('channel', channel)
    _multi_field('card_id', card_id)
    if min_amount is not None:
        where.append("amount >= ?")
        params.append(min_amount)
    if max_amount is not None:
        where.append("amount <= ?")
        params.append(max_amount)
    if start_ts:
        where.append("[timestamp] >= ?")
        params.append(_parse_ts(start_ts))
    if end_ts:
        where.append("[timestamp] <= ?")
        params.append(_parse_ts(end_ts))
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""

    # Count total
    total = 0
    with engine.connect() as conn:
        res = conn.exec_driver_sql(f"SELECT COUNT(1) FROM dbo.ht_Transactions {where_sql}", tuple(params))
        row = res.fetchone()
        total = int(row[0]) if row else 0

    # Page query
    offset = max(0, page) * max(1, page_size)
    items: List[Dict[str, Any]] = []
    # Validate sort column
    allowed_cols = {
        "txn_id": "txn_id",
        "employee_id": "employee_id",
        "merchant": "merchant",
        "city": "city",
        "category": "category",
        "amount": "amount",
        "timestamp": "[timestamp]",
        "channel": "channel",
        "card_id": "card_id",
    }
    order_col = allowed_cols.get((sort_by or "").lower(), "[timestamp]")
    order_dir = "ASC" if str(sort_dir).lower() == "asc" else "DESC"
    with engine.connect() as conn:
        # OFFSET-FETCH requires ORDER BY; allow simple column sort with tiebreaker txn_id
        q = (
            "SELECT txn_id, employee_id, merchant, city, category, amount, [timestamp], channel, card_id "
            f"FROM dbo.ht_Transactions {where_sql} ORDER BY {order_col} {order_dir}, txn_id OFFSET CAST(? AS INT) ROWS FETCH NEXT CAST(? AS INT) ROWS ONLY"
        )
        rows = conn.exec_driver_sql(q, tuple(params + [int(offset), int(page_size)]))
        cols = rows.keys()
        for r in rows:
            items.append({k: r[idx] for idx, k in enumerate(cols)})

    return {"items": items, "total": total, "page": page, "page_size": page_size}


def distinct_values(field: str, q: Optional[str] = None, limit: int = 50) -> List[str]:
    url = sqlalchemy_url_from_env()
    if not url:
        raise RuntimeError(
            "MSSQL environment variables not set. Set MSSQL_HOST, MSSQL_DB, MSSQL_USER, MSSQL_PASSWORD."
        )
    engine = create_engine_lazy(url)
    allowed = {"employee_id","merchant","city","category","channel","card_id"}
    if field not in allowed:
        raise ValueError(f"Unsupported field for distinct: {field}")
    where = ""
    params: List[Any] = []
    if q:
        where = f"WHERE {field} LIKE ?"
        params.append(f"%{q}%")
    sql = f"SELECT DISTINCT TOP {int(limit)} {field} FROM dbo.ht_Transactions {where} ORDER BY {field}"
    vals: List[str] = []
    with engine.connect() as conn:
        res = conn.exec_driver_sql(sql, tuple(params))
        for row in res:
            vals.append(str(row[0]))
    return vals

